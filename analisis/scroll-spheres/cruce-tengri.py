#!/usr/bin/env python3
"""Cruza el catálogo de @Scroll-Spheres con lo que ya ha adaptado @Tengri1337.

Tengri traduce a Scroll Spheres casi uno a uno, así que sus 39 shorts se
corresponden con 39 originales concretos. La correspondencia está escrita a
mano abajo —los títulos traducidos no coinciden literalmente, así que no hay
forma automática fiable de emparejarlos— y de ahí sale lo interesante: **qué
vídeos de Scroll Spheres funcionaron y Tengri todavía no ha tocado.**

    python3 analisis/scroll-spheres/cruce-tengri.py

Deja `hueco-vs-tengri.xlsx` con seis hojas: el plan de publicación, el catálogo
entero marcando en qué estado está cada vídeo, las candidatas libres que pasaron
del millón con su potencial, el rendimiento de cada adaptación de Tengri, el
ritmo mes a mes y el criterio de la criba. La criba vive en `potencial.py` y el
orden de publicación en `plan.py`.

Los colores del libro:
    verde  potencial alto, y libre
    gris   ya lo ha hecho Tengri
    azul   ya lo hemos hecho nosotros
"""
from __future__ import annotations

import datetime
import json
import statistics
import subprocess
import sys
from pathlib import Path

AQUI = Path(__file__).resolve().parent
sys.path.insert(0, str(AQUI))
import plan  # noqa: E402
import potencial as pot  # noqa: E402

# Título de Scroll Spheres ← short de Tengri que lo adapta.
# Se empareja por el título original porque los identificadores de Tengri no
# dicen nada sobre qué original están versionando.
ADAPTADOS = {
    "Insurance Companies Blacklisted Jackie Chan": "Jackie Chan Is Too Risky for Insurance Companies",
    "Tom Cruise Spent 5 Months Preparing For This Scene": "Tom Cruise Pasó 5 Meses Preparándose Para Esta Escena",
    "Ledger Ignored Nolan & Nailed It - The Dark Knight": "Ledger Improvised This Scene in One Take",
    "A $500,000 IMAX Camera Got Destroyed": "Catwoman Destruyó una Cámara de $500.000",
    "Jackie Chan Performs His Own Stunts - Rush Hour 3": "Jackie Chan Risked His Life For This Scene",
    "The Weapon That Defines Anton Chigurh": "The Terrifying Mindset of Anton Chigurh",
    "Kylo Ren Had The Most Terrifying Entrance": "Kylo Ren's Terrifying Entrance",
    "Jackie Chan Jumped Down 10 Stories At 52 Years Old": "Jackie Chan Rappelled Down 10 Floors Without CGI",
    "Christopher Lloyd Is Built Different - Nobody": "Christopher Lloyd Se Negó a Usar Armas Falsas",
    "Jason Statham Struggled With This Transporter 2 Scene": "Jason Statham Needed 12 Takes for This Scene",
    "The Narrator Did This To Get Rid of Tyler": "El Narrador Hizo Esto Para Deshacerse de Tyler",
    "He Finally Stared At The Baba Yaga - John Wick": "Finalmente Se Encontró Cara a Cara con Baba Yaga",
    "The Bullet Realism In John Wick: Chapter 3": "El Realismo de las Balas en John Wick: Capítulo 3",
    "The Realism Behind Ammo Tracking In John Wick 2": "El Realismo Detrás del Conteo de Balas en John Wick 2",
    "Gosling Faced His Fear For This Scene The Fall Guy": "Gosling Se Enfrentó a Su Miedo Para Esta Escena",
    "Why John Press Checks His Kimber - John Wick 2": "Why John Checks the Chamber of His Kimber",
    "The Ultimate John Wick Flex": "La Demostración de Poder Definitiva de John Wick",
    "Keanu Reeves' Stuntman Is A Legend - John Wick 3": "Keanu Reeves' Stunt Double Is a Legend",
    "The Coldest Flex In Movie History - John Wick 2": "La Demostración de Poder Más Brutal de la Historia del Cine",
    "He Knew To Leave Immediately - Nobody (2021)": "He Knew He Had to Leave Immediately",
    "Why Grace Reacts Like That In Project Hail Mary": "Why Grace Reacts That Way in Project Hail Mary",
    "The Toughest Henchman In John Wick 2": "El Secuaz Más Duro de John Wick 2",
    "The Marquis Was Doomed Either Way - John Wick 4": "The Marquis Was Doomed Either Way",
    "He Stopped The Gun With His Pinky - Nobody": "Detuvo el Arma con el Meñique",
    "Hunting Jaguar Paw On His Own Ground Was A Mistake": "Cazar a Jaguar Paw en Su Propio Territorio Fue un Error",
    "Vikings Mistook An Alien For A Dragon In Outlander": "Los Vikingos Confundieron a un Alienígena con un Dragón",
    "Apocalypto Is A Criminally Underrated Masterpiece": "Apocalypto Es una Obra Maestra Injustamente Infravalorada",
    "Nolan Pushed Anne Hathaway To Do Her Own Stunts": "Nolan Pushed Anne Hathaway to Do Her Own Stunts",
    "Keanu Called This Stuntman A Legend - John Wick": "Keanu Llamó a Este Especialista una Leyenda",
    "The Darkest Moment In Movie History - The Crow": "The Darkest Moment in Film History",
    "In Nobody He Strapped A Claymore To A Glass Shield": "En Nobody, Ató Una Mina Claymore a un Escudo de Cristal",
    "One Bullet Against An Ancient God - The Ritual": "Una Bala Contra un Dios Antiguo",
    # ── Adaptaciones del 20-25/08: Tengri también ha pasado a la inmediatez
    # y versiona lo nuevo de Scroll Spheres a los pocos días, a veces horas.
    "Nothing On Genna Is Harmless - Predator Badlands": "Este Gusano Explota Cuando Se Siente Amenazado",
    "The Real MVP In Predator: Badlands (2025)": "La Pequeña Criatura Que Protegió a Dek Hasta el Final",
    "First Time Dek Showed Fear In Predator: Badlands": "El Momento en Que Dek Sintió Miedo de Verdad",
    "The Only Creature That Can Annoy A Predator": "La Única Criatura Capaz de Molestar a un Depredador",
    "Movies Overexaggerate About How Tough We Are": "Una Sola Mala Caída Puede Acabar Con Todo",
    "She Couldn't Breathe - The Descent": "Una Sola Respiración Podía Costarle la Vida",
    "James McAvoy’s Acting Masterclass In Glass (2019)": "James McAvoy Played 24 Personalities",
    "Mark Wahlberg Caught Denzel Off Guard - 2 Guns": "Mark Wahlberg Improvised This Scene",
    "Hugh Jackman's Reaction Was Hilarious": "Hugh Jackman Made Rob Delaney Break Character",
    "It Was Over Once He Touched Water - Aquaman (2023)": "They Chased Him to His Own Source of Power",
    "The Western Tribute Hidden In John Wick": "John Wick Recreó Esta Icónica Escena del Western",
    "Everybodys Gangsta Until The Stopwatch Equalizer 3": "They Only Had 9 Seconds to Choose",
    "The Maze Runner Death Cure Dylan O Brien's Stunt": "Dylan O’Brien Casi Muere Filmando Esta Escena",
    "Batman Part II Behind The Scenes Footage": "El SWAT Abrió Fuego Contra el Batmóvil",
    "It Follows (2014) Came From A Real Nightmare": "This Monster Was Born from a Real Nightmare",
    "They Ran So Differently In Insurgent (2015) 💀": "The Actors Interpreted This Instruction in Two Very Different Ways",
    # Scroll Spheres tiene dos shorts de Project Power. Tengri habla de un
    # impacto concreto, así que apunta a este y no al general de 10 M.
    "That Face Ripple Was 100% Practical Project Power": "Este Impacto de Bala Fue Real — Project Power",
}

# Shorts de Tengri cuyo original no he sabido localizar en el catálogo. Se
# listan para no dar por libre un tema que quizá ya esté cogido.
SIN_LOCALIZAR = [
    "I Didn't Have a Plan B",
    "Failed to Defuse the Nuclear Bomb — Fallout",
    "Este Detalle Reveló Que No Era un Policía de Verdad",
    "10 Años Después, Sigue Siendo el Mejor Deslizamiento del Cine — Maze Runner",
    "Tom Cruise Confió Su Vida a Hiroyuki Sanada — The Last Samurai",
    "Keanu Reeves Cobraba $39.000 Por Palabra — John Wick 4",
    # nuevos del 25/08 sin original claro en el catálogo
    "Tom Cruise Cambió de Disfraz en Plena Toma — Ghost Protocol",
    "Watchmen No Era una Película Familiar — Watchmen (2009)",
    "Jonathan Bailey Had an Embarrassing Moment in Front of Scarlett Johansson",
    "Nicolas Cage Se Entregó por Completo en Esta Escena",
    "Heath Ledger Never Broke Character in This Scene — The Dark Knight",
    "He Caught the Spear and Threw It Back — Apocalypto",
    "Krypto Bit Mr. Terrific With Superman's Strength — Superman (2025)",
    "Este Detalle Reveló el Destino de Eddie Brock — Spider-Man 3 (¿el mismo que Eddie Vaporized? comprobar)",
]


# Texto de la hoja «Criterio». (línea, en negrita)
CRITERIO = [
    ("Cómo se ha hecho la criba", True),
    ("", False),
    ("Tengri traduce a Scroll Spheres uno a uno, así que sus 39 shorts son un "
     "experimento ya pagado: sabemos qué original rindió y cuál no al pasarlo "
     "al castellano. El verde sale de ahí, no de una intuición.", False),
    ("", False),
    ("Las cuatro preguntas", True),
    ("1. ¿Se entiende en los tres primeros segundos, sin haber visto la "
     "película?", False),
    ("2. ¿Hay algo físico que ver, o es un dato que hay que contar?", False),
    ("3. ¿Sobrevive a la traducción? Los acentos, los juegos de palabras y las "
     "frases míticas en inglés se mueren en castellano, y más con el doblaje.", False),
    ("4. ¿La película o el actor significan algo en España?", False),
    ("", False),
    ("Las cuatro a favor es «alto» y va en verde. Una en contra, «medio». Dos o "
     "más, «bajo».", False),
    ("", False),
    ("Por qué no basta con ordenar por visitas", True),
    ("De los shorts que Tengri adaptó, los tres originales de 39 millones le "
     "retuvieron un 2,6 %, un 0,7 % y un 0,6 %. En cambio Transporter 2, que en "
     "Scroll Spheres se quedó en 26.000 visitas, le hizo 468.000; y Anton "
     "Chigurh, con 43.000 de origen, le hizo 352.000. La mediana de conversión "
     "es del 1,9 %, pero el reparto no guarda relación con el tamaño del "
     "original.", False),
    ("", False),
    ("Con una salvedad honesta: parte de sus cifras bajas son de shorts recién "
     "publicados. Aun así, entre los más antiguos hay tanto 4,7 millones como "
     "1.400 visitas, así que la varianza es real y no solo cuestión de "
     "antigüedad.", False),
    ("", False),
    ("El corte del millón", True),
    ("De los 1.758 shorts del canal, 1.476 no llegan a 100.000 visitas y la "
     "mediana está en 25.127. Los 100 mejores se llevan el 87 % de los mil "
     "millones de visitas. Por debajo del millón hay muy poco que rascar, y por "
     "eso la hoja de candidatas corta ahí.", False),
    ("", False),
    ("Con las fechas ya puestas se ha comprobado: ordenando por «× mediana del "
     "mes» —que corrige la ventaja de lo antiguo y el crecimiento del canal— "
     "los 60 primeros son los mismos vídeos que ya estaban por encima del "
     "millón. Ninguno se quedaba fuera por el corte.", False),
    ("", False),
    ("Lo que dicen las fechas", True),
    ("1.055.801.922 visitas en 362 días, del 24/08/2025 al 20/08/2026.", False),
    ("", False),
    ("El canal subió el ritmo de 1,9 shorts al día hasta 10 al día en enero de "
     "2026, y desde entonces ha bajado a 3,8. Lo interesante es la mediana de "
     "visitas en ese recorrido: 31.785 en enero, 11.469 en abril —el peor mes, "
     "y el de más volumen acumulado— y 38.083 en julio. Publicando menos de la "
     "mitad, su suelo se ha triplicado. El propio Scroll Spheres ha abandonado "
     "la estrategia de volumen.", False),
    ("", False),
    ("Un short madura rápido: la mayoría de sus visitas llegan en la primera "
     "semana. A los siete días ya se puede juzgar si ha funcionado, no hace "
     "falta esperar un mes.", False),
    ("", False),
    ("Colores", True),
    ("verde: libre y de potencial alto — por aquí se empieza", False),
    ("gris: lo tiene Tengri — se puede hacer igual, pero compites de frente", False),
    ("azul: ya lo hemos preparado nosotros", False),
    ("sin color: libre, de potencial medio o bajo", False),
]


def tengri() -> dict[str, int]:
    """Visitas de cada short de Tengri, por título."""
    r = subprocess.run(
        ["yt-dlp", "--flat-playlist", "-J", "--no-warnings",
         "https://www.youtube.com/@Tengri1337/shorts"],
        capture_output=True, text=True, timeout=900)
    if r.returncode != 0:
        sys.exit(f"yt-dlp ha fallado:\n{r.stderr.strip()[-500:]}")
    return {e["title"]: e.get("view_count") or 0
            for e in json.loads(r.stdout).get("entries") or []}


def fechas_y_visitas(ss: list[dict]) -> str:
    """Mete la fecha y las visitas exactas del csv, si ya se han conseguido.

    El catálogo de yt-dlp trae las visitas redondeadas y ninguna fecha; el csv
    lo genera `shorts-canal.py` leyendo la ficha de cada vídeo, que solo se
    puede hacer desde una conexión doméstica.
    """
    csv_ruta = AQUI / "shorts-scrollspheres.csv"
    if not csv_ruta.exists():
        return "sin fechas: falta shorts-scrollspheres.csv"

    import csv as _csv
    with open(csv_ruta, encoding="utf-8-sig") as f:
        filas = {r["Título"]: r for r in _csv.DictReader(f, delimiter=";")}

    puestas = 0
    for v in ss:
        r = filas.get(v["titulo"])
        if not r or not r.get("Fecha de subida"):
            continue
        d, m, a = r["Fecha de subida"].split("/")
        v["fecha"] = datetime.date(int(a), int(m), int(d))
        v["vistas"] = int(r["Visitas"])
        puestas += 1

    # x mediana: visitas partido por la mediana de su mes de publicación. Sin
    # esto el ranking premia lo antiguo, que ha tenido más tiempo de acumular,
    # y castiga lo reciente aunque vaya mucho mejor.
    por_mes: dict[str, list[int]] = {}
    for v in ss:
        if v.get("fecha"):
            por_mes.setdefault(v["fecha"].strftime("%Y-%m"), []).append(v["vistas"])
    medianas = {k: statistics.median(g) for k, g in por_mes.items()}
    for v in ss:
        if v.get("fecha"):
            v["x_mes"] = v["vistas"] / medianas[v["fecha"].strftime("%Y-%m")]
    return f"{puestas}/{len(ss)} con fecha y visitas exactas"


def main() -> None:
    ss = json.loads((AQUI / "shorts-scrollspheres-catalogo.json").read_text("utf8"))
    for v in ss:
        v["fecha"], v["x_mes"] = None, None
    nota_fechas = fechas_y_visitas(ss)
    tg = tengri()

    # Los títulos de Tengri en ADAPTADOS son un prefijo del real (llevan
    # « — Película» detrás), así que se busca por comienzo de cadena.
    def visitas_tengri(titulo_tg: str) -> int | None:
        for t, v in tg.items():
            if t.startswith(titulo_tg[:40]):
                return v
        return None

    for v in ss:
        adapt = ADAPTADOS.get(v["titulo"])
        v["tengri"] = "sí" if adapt else ""
        v["tengri_vistas"] = visitas_tengri(adapt) if adapt else None
        tuyo = v["id"] in pot.TUYOS
        v["estado"] = ("ya es tuyo (y de Tengri)" if tuyo and adapt else
                       "ya es tuyo" if tuyo else
                       "lo tiene Tengri" if adapt else "libre")
        v["potencial"], v["porque"] = pot.POTENCIAL.get(v["titulo"], ("", ""))

    ss.sort(key=lambda v: v["vistas"] or 0, reverse=True)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    verde = PatternFill("solid", fgColor="C6EFCE")   # potencial alto
    gris = PatternFill("solid", fgColor="DDDDDD")    # ya cogido
    azul = PatternFill("solid", fgColor="DDEBF7")    # ya es tuyo

    def enlazar(hoja, fila, col, vid):
        """Deja la URL completa y pinchable. Es la que se copia para pedir la
        portada, los títulos, la descripción y las etiquetas."""
        if not vid:
            return
        celda = hoja.cell(fila, col)
        celda.value = f"https://www.youtube.com/shorts/{vid}"
        celda.hyperlink = celda.value
        celda.font = Font(color="0563C1", underline="single")
        celda.alignment = Alignment(vertical="top")

    def pintar(hoja, v):
        relleno = (azul if v["estado"].startswith("ya es tuyo") else
                   gris if v["estado"] == "lo tiene Tengri" else
                   verde if v["potencial"] == "alto" else None)
        if relleno:
            for c in hoja[hoja.max_row]:
                c.fill = relleno

    wb = Workbook()
    ws = wb.active
    ws.title = "Hueco"
    ws.append(["#", "Título (Scroll Spheres)", "Enlace", "Visitas SS", "Fecha",
               "× mediana del mes", "Estado", "Visitas Tengri", "Potencial", "Por qué"])
    for c in ws[1]:
        c.font = Font(bold=True)

    for i, v in enumerate(ss, 1):
        ws.append([i, v["titulo"], None, v["vistas"], v["fecha"], v["x_mes"],
                   v["estado"], v["tengri_vistas"], v["potencial"], v["porque"]])
        pintar(ws, v)
        enlazar(ws, ws.max_row, 3, v.get("id"))

    for col, ancho in ((1, 6), (2, 66), (3, 45), (4, 13), (5, 12), (6, 17),
                       (7, 16), (8, 14), (9, 11), (10, 92)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for fila in ws.iter_rows(min_row=2, min_col=4, max_col=9):
        fila[0].number_format = fila[4].number_format = "#,##0"
        fila[1].number_format = "DD/MM/YYYY"
        fila[2].number_format = "0"
        for j in (1, 2, 3, 5):
            fila[j].alignment = Alignment(horizontal="center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    cogidos = [v for v in ss if v["tengri"]]
    libres = [v for v in ss if v["estado"] == "libre"]

    # Segunda hoja: lo que de verdad hay que mirar. El catálogo entero son
    # 1.758 filas de las que 1.476 no llegan a 100.000 visitas; la cola larga
    # no dice nada.
    # Ordenados por potencial primero y por visitas después: la lista se lee
    # de arriba abajo y lo verde queda junto.
    orden = {"alto": 0, "medio": 1, "bajo": 2, "": 3}
    candidatas = sorted([x for x in libres if (x["vistas"] or 0) >= 1_000_000],
                        key=lambda v: (orden[v["potencial"]], -(v["vistas"] or 0)))

    h2 = wb.create_sheet("Libres +1M")
    h2.append(["#", "Título (Scroll Spheres)", "Enlace", "Visitas SS", "Fecha",
               "× mediana del mes", "Potencial", "Por qué"])
    for c in h2[1]:
        c.font = Font(bold=True)
    for i, v in enumerate(candidatas, 1):
        h2.append([i, v["titulo"], None, v["vistas"], v["fecha"], v["x_mes"],
                   v["potencial"], v["porque"]])
        pintar(h2, v)
        enlazar(h2, h2.max_row, 3, v.get("id"))
    for col, ancho in ((1, 6), (2, 66), (3, 45), (4, 13), (5, 12), (6, 17),
                       (7, 11), (8, 96)):
        h2.column_dimensions[get_column_letter(col)].width = ancho
    for fila in h2.iter_rows(min_row=2, min_col=4, max_col=7):
        fila[0].number_format = "#,##0"
        fila[1].number_format = "DD/MM/YYYY"
        fila[2].number_format = "0"
        for j in (1, 2, 3):
            fila[j].alignment = Alignment(horizontal="center")
    h2.freeze_panes = "C2"
    h2.auto_filter.ref = f"A1:H{h2.max_row}"

    # Tercera hoja: qué le rindió a Tengri cada adaptación. Es lo único que
    # dice de verdad cuánto se traduce una cifra inglesa al castellano.
    h3 = wb.create_sheet("Rendimiento Tengri")
    h3.append(["Título (Scroll Spheres)", "Visitas SS", "Visitas Tengri", "% que retiene"])
    for c in h3[1]:
        c.font = Font(bold=True)
    for v in sorted(cogidos, key=lambda x: -(x["tengri_vistas"] or 0)):
        a, b = v["vistas"] or 0, v["tengri_vistas"]
        h3.append([v["titulo"], a, b, (b / a) if (a and b is not None) else None])
    for col, ancho in ((1, 70), (2, 13), (3, 14), (4, 14)):
        h3.column_dimensions[get_column_letter(col)].width = ancho
    for fila in h3.iter_rows(min_row=2, min_col=2, max_col=4):
        fila[0].number_format = fila[1].number_format = "#,##0"
        fila[2].number_format = "0,0%"
    h3.freeze_panes = "A2"

    # Primera hoja del libro: qué hacer, en qué orden y por qué. Va delante
    # porque es lo único que hay que leer para ponerse a trabajar.
    hp = wb.create_sheet("Plan", 0)
    hp.append(["Orden", "Día", "Título (Scroll Spheres)", "Enlace al original",
               "Visitas SS", "× mediana del mes", "Por qué en esta posición"])
    for c in hp[1]:
        c.font = Font(bold=True)

    # El amarillo lo puso el usuario a mano en su copia para marcar lo ya
    # publicado; aquí se respeta ese color para que su marca no se pierda al
    # regenerar. Los días se cuentan solo sobre lo pendiente, desde hoy.
    amarillo = PatternFill("solid", fgColor="FFFF00")
    por_titulo = {v["titulo"]: v for v in ss}
    pendientes = 0
    for i, (titulo, motivo) in enumerate(plan.ORDEN, 1):
        v = por_titulo.get(titulo, {})
        publicado = titulo in plan.PUBLICADOS
        if publicado:
            dia = "✔"
        else:
            pendientes += 1
            dia = (pendientes - 1) // plan.RITMO + 1
        hp.append([i, dia, titulo, None, v.get("vistas"), v.get("x_mes"), motivo])
        for c in hp[hp.max_row]:
            c.fill = amarillo if publicado else verde
        enlazar(hp, hp.max_row, 4, v.get("id"))
    for col, ancho in ((1, 7), (2, 6), (3, 66), (4, 45), (5, 13), (6, 17), (7, 104)):
        hp.column_dimensions[get_column_letter(col)].width = ancho
    for fila in hp.iter_rows(min_row=2, min_col=1, max_col=7):
        fila[4].number_format = "#,##0"
        fila[5].number_format = "0"
        for j in (0, 1, 4, 5):
            fila[j].alignment = Alignment(horizontal="center")
        fila[6].alignment = Alignment(wrap_text=True, vertical="top")
        for j in (2, 3):
            fila[j].alignment = Alignment(vertical="top")
    hp.freeze_panes = "C2"

    # El criterio del plan, debajo de la tabla, para que viaje con ella. El
    # docstring lleva marcas de markdown y viñetas propias: en una celda
    # estorban, así que se limpian y cada viñeta va en su fila.
    hp.append([])
    for bloque in plan.__doc__.split("\n\n")[1:]:
        for n, trozo in enumerate(bloque.split("\n  · ")):
            texto = " ".join(trozo.split()).replace("**", "").replace("`", "")
            if not texto:
                continue
            if n and not texto.startswith("· "):
                texto = "· " + texto
            hp.append(["", "", texto])
            hp.cell(hp.max_row, 3).alignment = Alignment(wrap_text=True,
                                                         vertical="top")

    # Cuarta hoja: el ritmo mes a mes. Es donde se ve que el canal abandonó el
    # volumen, que es lo más útil que dicen las fechas.
    con_fecha = [v for v in ss if v.get("fecha")]
    if con_fecha:
        meses: dict[str, list[dict]] = {}
        for v in sorted(con_fecha, key=lambda v: v["fecha"]):
            meses.setdefault(v["fecha"].strftime("%Y-%m"), []).append(v)

        hr = wb.create_sheet("Ritmo")
        hr.append(["Mes", "Shorts", "Al día", "Mediana de visitas",
                   "Visitas del mes", "El mejor del mes"])
        for c in hr[1]:
            c.font = Font(bold=True)
        for k, g in meses.items():
            vs = [x["vistas"] for x in g]
            dias = len({x["fecha"] for x in g})
            hr.append([k, len(g), round(len(g) / dias, 1),
                       int(statistics.median(vs)), sum(vs), max(vs)])
        for col, ancho in ((1, 10), (2, 9), (3, 9), (4, 20), (5, 18), (6, 18)):
            hr.column_dimensions[get_column_letter(col)].width = ancho
        for fila in hr.iter_rows(min_row=2, min_col=2, max_col=6):
            for j in (0, 2, 3, 4):
                fila[j].number_format = "#,##0"
            fila[1].number_format = "0.0"
        hr.freeze_panes = "A2"

    # Quinta hoja: de dónde sale el verde, para que la criba se pueda discutir
    # en vez de tener que creérsela.
    h4 = wb.create_sheet("Criterio")
    h4.column_dimensions["A"].width = 118
    for linea, negrita in CRITERIO:
        h4.append([linea])
        if negrita:
            h4[h4.max_row][0].font = Font(bold=True)
        h4[h4.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")

    for hoja, _ in ((ws, None), (h2, None)):
        for fila in hoja.iter_rows(min_row=2, min_col=hoja.max_column,
                                   max_col=hoja.max_column):
            fila[0].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(AQUI / "hueco-vs-tengri.xlsx")

    publicados = [t for t, _ in plan.ORDEN if t in plan.PUBLICADOS]
    quedan = len(plan.ORDEN) - len(publicados)
    print(f"  plan: {len(publicados)} publicados (en amarillo) · {quedan} pendientes "
          f"(~{-(-quedan // plan.RITMO)} días a {plan.RITMO}/día)")
    altos = [v for v in candidatas if v["potencial"] == "alto"]
    print(f"✓ hueco-vs-tengri.xlsx · {len(ss)} shorts de Scroll Spheres")
    print(f"  {len(cogidos)} los tiene Tengri · {len(pot.TUYOS)} ya son tuyos "
          f"· {len(libres)} libres")
    print(f"  sin localizar el original: {len(SIN_LOCALIZAR)} shorts de Tengri")
    print(f"  {nota_fechas}")
    print(f"  candidatas libres con +1M: {len(candidatas)}, "
          f"de las que {len(altos)} en verde")
    for corte in (10_000_000, 5_000_000, 1_000_000):
        n = sum(1 for v in libres if (v["vistas"] or 0) >= corte)
        print(f"  libres con más de {corte:,} visitas: {n}".replace(",", "."))


if __name__ == "__main__":
    main()
