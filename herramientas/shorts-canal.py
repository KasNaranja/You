#!/usr/bin/env python3
"""Lista todos los Shorts de un canal de YouTube en una hoja de cálculo.

Tres columnas: título, visitas y fecha de subida en DD/MM/AAAA.

    python3 herramientas/shorts-canal.py https://www.youtube.com/@Scroll-Spheres \
        --salida shorts-scrollspheres --clave AIza…

El catálogo (los identificadores y los títulos) sale siempre de yt-dlp, que
recorre la pestaña /shorts entera. Lo que no da esa pestaña es la fecha: el
mosaico de Shorts solo muestra el contador de visitas redondeado, ni siquiera un
«hace 3 semanas». Para la fecha hay dos vías:

  · `--clave`: la API oficial de YouTube (`videos.list`). 50 vídeos por
    petición, fecha y visitas exactas. Para 1.758 shorts son 36 peticiones y
    36 unidades de cuota de las 10.000 diarias que da una clave gratuita.

  · sin clave: se lee la página de cada vídeo, que lleva la fecha exacta.
    Es una petición por vídeo y **solo funciona desde una conexión doméstica**:
    desde un centro de datos Google devuelve un 429 con captcha a las pocas
    decenas de peticiones. Guarda el avance en `<salida>-avance.json`, así que
    se puede cortar y reanudar.

La API interna (`youtubei/v1/player`), que sería lo más ligero, responde
LOGIN_REQUIRED desde cualquier IP de centro de datos.
"""
from __future__ import annotations

import argparse
import csv
import gzip
import json
import re
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

CABECERAS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"),
    # En inglés la fecha sale siempre como «Aug 20, 2026»; en otros idiomas
    # cambia el formato y habría que mantener una tabla por idioma.
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip",
}

RE_FECHA = re.compile(r'"dateText":\{"simpleText":"([^"]+)"')
RE_VISTAS = re.compile(r'"viewCountNumber":"(\d+)"')
RE_VISTAS_ALT = re.compile(r'"viewCount":\{"simpleText":"([\d.,]+) views?"')

MESES = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
         "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}


# ── catálogo ──────────────────────────────────────────────────────────────
def catalogo(canal: str) -> list[dict]:
    """Lista plana de la pestaña /shorts, vía yt-dlp."""
    url = canal.rstrip("/")
    if not url.endswith("/shorts"):
        url += "/shorts"
    print(f"Leyendo el catálogo de {url} …", flush=True)
    r = subprocess.run(["yt-dlp", "--flat-playlist", "-J", "--no-warnings", url],
                       capture_output=True, text=True, timeout=1800)
    if r.returncode != 0:
        sys.exit(f"yt-dlp ha fallado:\n{r.stderr.strip()[-800:]}")
    datos = json.loads(r.stdout)

    vistos, lista = set(), []
    for e in datos.get("entries") or []:
        vid = e.get("id")
        if not vid or vid in vistos:
            continue
        vistos.add(vid)
        lista.append({"id": vid, "titulo": e.get("title") or "",
                      "vistas": e.get("view_count"), "fecha": ""})
    print(f"  {len(lista)} shorts en el canal.", flush=True)
    return lista


def bajar(url: str, timeout: int = 45) -> str:
    r = urllib.request.urlopen(
        urllib.request.Request(url, headers=CABECERAS), timeout=timeout)
    crudo = r.read()
    if r.headers.get("Content-Encoding") == "gzip":
        crudo = gzip.decompress(crudo)
    return crudo.decode("utf8", "ignore")


def a_ddmmaaaa(texto: str) -> str:
    """«Aug 20, 2026» → «20/08/2026». También acepta «Premiered Aug 20, 2026»."""
    m = re.search(r"([A-Z][a-z]{2})\s+(\d{1,2}),\s*(\d{4})", texto)
    if not m:
        return ""
    mes = MESES.get(m.group(1))
    return f"{int(m.group(2)):02d}/{mes:02d}/{m.group(3)}" if mes else ""


# ── vía 1: API oficial ────────────────────────────────────────────────────
def por_api(lista: list[dict], clave: str) -> list[dict]:
    """`videos.list` en tandas de 50: fecha y visitas exactas."""
    porid = {v["id"]: v for v in lista}
    ids = list(porid)
    print(f"Consultando la API oficial ({-(-len(ids) // 50)} peticiones)…",
          flush=True)

    for i in range(0, len(ids), 50):
        tanda = ids[i:i + 50]
        url = ("https://www.googleapis.com/youtube/v3/videos?part=snippet,statistics"
               f"&maxResults=50&id={','.join(tanda)}&key={urllib.parse.quote(clave)}")
        try:
            d = json.loads(bajar(url))
        except urllib.error.HTTPError as e:
            detalle = e.read().decode("utf8", "ignore")[:400]
            sys.exit(f"La API ha respondido {e.code}:\n{detalle}")

        for it in d.get("items", []):
            v = porid.get(it["id"])
            if not v:
                continue
            v["titulo"] = it["snippet"]["title"]
            # publishedAt viene en UTC: 2026-08-20T14:03:12Z
            a, m, dia = it["snippet"]["publishedAt"][:10].split("-")
            v["fecha"] = f"{dia}/{m}/{a}"
            cuenta = it.get("statistics", {}).get("viewCount")
            if cuenta is not None:
                v["vistas"] = int(cuenta)
        hechos = min(i + 50, len(ids))
        print(f"  {hechos}/{len(ids)}", flush=True)

    sin = [v for v in lista if not v["fecha"]]
    if sin:
        # Un vídeo borrado o privado entre el listado y la consulta.
        print(f"  ! {len(sin)} sin respuesta de la API: "
              + ", ".join(v["id"] for v in sin[:10]))
    return lista


# ── vía 2: raspado de la página de cada vídeo ─────────────────────────────
def por_paginas(lista: list[dict], pausa: float, avance: Path) -> list[dict]:
    """Una petición por vídeo. Lento y solo viable desde conexión doméstica."""
    cache: dict[str, dict] = {}
    if avance.exists():
        cache = json.loads(avance.read_text(encoding="utf8"))
        print(f"Reanudando: {len(cache)} vídeos ya resueltos.", flush=True)

    pendientes = [v for v in lista if v["id"] not in cache]
    print(f"Leyendo la página de {len(pendientes)} vídeos "
          f"(pausa de {pausa}s entre peticiones)…", flush=True)

    t0, castigo = time.time(), 0.0
    for n, v in enumerate(pendientes, 1):
        for intento in range(4):
            try:
                h = bajar(f"https://www.youtube.com/watch?v={v['id']}&hl=en")
                castigo = max(0.0, castigo - 0.5)
                break
            except urllib.error.HTTPError as e:
                if e.code == 429:
                    # Captcha por IP. Cada vez se espera más; si ni así,
                    # se para: seguir insistiendo solo alarga el bloqueo.
                    castigo += 2
                    espera = 30 * (intento + 1)
                    print(f"  429 en {v['id']}: esperando {espera}s…", flush=True)
                    time.sleep(espera)
                else:
                    time.sleep(2 ** intento)
            except Exception:
                time.sleep(2 ** intento)
        else:
            _guardar_avance(avance, cache)
            print(f"\n! Bloqueo persistente tras {len(cache)} vídeos. "
                  f"El avance está en {avance}: se puede reanudar más tarde "
                  f"o terminar con --clave.")
            break

        f = RE_FECHA.search(h)
        m = RE_VISTAS.search(h) or RE_VISTAS_ALT.search(h)
        cache[v["id"]] = {
            "fecha": a_ddmmaaaa(f.group(1)) if f else "",
            "vistas": int(re.sub(r"\D", "", m.group(1))) if m else v["vistas"],
        }

        if n % 25 == 0:
            _guardar_avance(avance, cache)
            seg = time.time() - t0
            queda = seg / n * (len(pendientes) - n)
            print(f"  {n}/{len(pendientes)} · {seg / 60:.0f} min "
                  f"· quedan ~{queda / 60:.0f} min", flush=True)
        time.sleep(pausa + castigo)

    _guardar_avance(avance, cache)
    for v in lista:
        if v["id"] in cache:
            v.update(cache[v["id"]])
    return lista


def _guardar_avance(ruta: Path, cache: dict) -> None:
    ruta.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf8")


# ── salida ────────────────────────────────────────────────────────────────
def clave_fecha(v: dict) -> tuple:
    """Ordena por fecha real, no por el texto DD/MM/AAAA."""
    f = v.get("fecha") or ""
    if len(f) == 10:
        d, m, a = f.split("/")
        return (int(a), int(m), int(d))
    return (0, 0, 0)


def guardar_xlsx(lista: list[dict], ruta: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Shorts"
    ws.append(["Título", "Visitas", "Fecha de subida"])
    for c in ws[1]:
        c.font = Font(bold=True)

    for v in lista:
        ws.append([v.get("titulo", ""), v.get("vistas"), v.get("fecha", "")])

    for col, ancho in ((1, 78), (2, 14), (3, 18)):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    for fila in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        fila[0].number_format = "#,##0"
        fila[1].alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    wb.save(ruta)


def guardar_csv(lista: list[dict], ruta: str) -> None:
    # utf-8-sig y punto y coma: así Excel en español lo abre en columnas
    # y no destroza las tildes.
    with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Título", "Visitas", "Fecha de subida"])
        for v in lista:
            w.writerow([v.get("titulo", ""), v.get("vistas", ""), v.get("fecha", "")])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("canal", help="URL o handle del canal")
    ap.add_argument("--salida", default="shorts", help="nombre base de los ficheros")
    ap.add_argument("--clave", help="clave de la API de YouTube (recomendado)")
    ap.add_argument("--pausa", type=float, default=1.5,
                    help="segundos entre peticiones al raspar sin clave")
    ap.add_argument("--catalogo", help="reutilizar un catálogo ya descargado (.json)")
    args = ap.parse_args()

    if args.catalogo:
        lista = json.loads(Path(args.catalogo).read_text(encoding="utf8"))
    else:
        lista = catalogo(args.canal)
        Path(f"{args.salida}-catalogo.json").write_text(
            json.dumps(lista, ensure_ascii=False), encoding="utf8")
    if not lista:
        sys.exit("El canal no ha devuelto ningún short.")

    if args.clave:
        lista = por_api(lista, args.clave)
    else:
        lista = por_paginas(lista, args.pausa, Path(f"{args.salida}-avance.json"))

    lista.sort(key=clave_fecha, reverse=True)
    guardar_xlsx(lista, f"{args.salida}.xlsx")
    guardar_csv(lista, f"{args.salida}.csv")

    con_fecha = [v for v in lista if v.get("fecha")]
    total = sum(v["vistas"] for v in lista if isinstance(v.get("vistas"), int))
    print(f"\n✓ {args.salida}.xlsx y {args.salida}.csv · {len(lista)} shorts")
    print(f"  {len(con_fecha)} con fecha", end="")
    if con_fecha:
        print(f", del {con_fecha[-1]['fecha']} al {con_fecha[0]['fecha']}", end="")
    print(f" · {total:,}".replace(",", ".") + " visitas en total")


if __name__ == "__main__":
    main()
