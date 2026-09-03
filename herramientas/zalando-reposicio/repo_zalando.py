# -*- coding: utf-8 -*-
"""
Càlcul de reposició Zalando per SKU i per model_color.

Fonts (carpeta de dades, per defecte la carpeta on hi ha aquest script):
  Models a reposar.xlsx              llista base de SKUs a reposar (EAN, SKU, season, gènere, model_color, talla...)
  NIVEL.xlsx                         nivells i desglossament per talla (MUJER / CABALLERO / NIÑO)
  Venda 2025.xlsx                    acumulat 2025 per model_color
  Vendes 2026/<mes>/VENDES DEL dd.mm al dd.mm.xlsx   vendes setmanals (pestanya DADES2, una línia per comanda)
  Stock Toni Pons/*.txt              export SAP (UTF-16, tabuladors): Stock 01 02, Disponible 30 / 59 dies
  Stock Zalando/*.csv|*.xlsx         stock snapshot Zalando (EAN, Offerable, Non-offerable, Total)
  Enviaments pendents/*.csv          enviaments ja fets però encara no al snapshot (ean;quantity)
  Ajustos repo.xlsx (opcional)       multiplicador / nivell forçat per model_color

Regla:
  objectiu = venda setmanal del model_color x MULT (2 per defecte)
  nivell   = el primer nivell de la taula del gènere el TOTAL DE PARELLS del qual (suma de totes les talles)
             cobreix l'objectiu (p.ex. MUJER 166 parells -> nivell 35, que en suma 188)
  HAURIA   = desglossament per talla d'aquest nivell
  DIF      = HAURIA - stock Zalando (total) - enviaments pendents
  REPO     = DIF si és positiu
  PREPARABLE = min(REPO, stock disponible 59 dies a Toni Pons)

Sortides (carpeta de sortida, per defecte la mateixa):
  Venda 2026 dd.mm.xlsx              consolidat de totes les setmanes (un fitxer per data de càlcul; el bo és sempre l'últim)
  REPO ZALANDO dd.mm.xlsx            pestanyes CÀLCUL SKU, MODEL_COLOR, FORA LLISTA, PARÀMETRES, NIVELLS
  REPO ZALANDO dd.mm.html            mateixes dues vistes, filtrables i ordenables

Ús:
  python repo_zalando.py                       (data = avui, MULT = 2, sense nivell mínim)
  python repo_zalando.py --mult 2 --min-level 6 --min-level-kids 2
  python repo_zalando.py --data "C:\\...\\Zalando reposició" --out "C:\\...\\sortida"
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import html
import json
import os
import re
import sys
import unicodedata

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------------------
# Paràmetres
# --------------------------------------------------------------------------------------
GENDER_GROUP = {  # GÈNERE de "Models a reposar" -> bloc de NIVEL.xlsx
    "DONA": "MUJER",
    "HOME": "CABALLERO",
    "NENS": "NIÑO",
    "UNISEX": "MUJER",       # taula de dona; talles 46-47 = valor de la talla 45
    "MINI": "NIÑO",          # taula de nen; talles < 25 = valor de la talla 25
    "COMPLEMENTS": None,     # sense taula: HAURIA = objectiu (talla única)
}
KIDS_GROUPS = {"NIÑO"}
WEEK_RE = re.compile(r"DEL (\d\d)\.(\d\d) al (\d\d)\.(\d\d)", re.I)
SNAP_DATE_RE = re.compile(r"(\d\d)[_\-.](\d\d)[_\-.](\d{4})")
EXCLUDE_SALES = ("acumulat", "càlcul", "calcul", "ranking", "anàlisi", "analisi")
HTML_HIDE = {"VENDA 4 SETM"}  # columnes de l'Excel que no es mostren a l'HTML


def norm(s: str) -> str:
    """minúscules sense accents, per comparar capçaleres."""
    s = unicodedata.normalize("NFKD", str(s))
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()


def clean_ean(x) -> str | None:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    if not re.fullmatch(r"\d{8,14}", s) or int(s) == 0:
        return None
    return s


def to_num(series: pd.Series) -> pd.Series:
    """'1.234,00' -> 1234.0 ; '12,00' -> 12.0 ; números tal qual."""
    if series.dtype.kind in "if":
        return series.fillna(0)
    s = series.astype(str).str.strip()
    s = s.where(~s.isin(["", "nan", "None"]), "0")
    has_comma = s.str.contains(",", regex=False)
    s = s.where(~has_comma, s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False))
    return pd.to_numeric(s, errors="coerce").fillna(0)


# --------------------------------------------------------------------------------------
# Càrrega de fonts
# --------------------------------------------------------------------------------------
def load_models(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype={"codigo_barra": str, "SKU": str})
    cols = {norm(c): c for c in df.columns}
    ren = {}
    for want, keys in {
        "EAN": ["codigo_barra", "ean"],
        "SKU": ["sku"],
        "SEASON": ["season"],
        "TEMPORADA": ["temporada"],
        "COL·LECCIÓ": ["col·leccio", "colleccio", "coleccion", "col·lecció"],
        "GÈNERE": ["genere", "gènere", "sexe", "genero"],
        "model": ["model"],
        "color": ["color"],
        "model_color": ["model_color"],
        "talla": ["talla"],
        "ES POT ENVIAR?": ["es pot enviar?", "es pot enviar"],
    }.items():
        for k in keys:
            if norm(k) in cols:
                ren[cols[norm(k)]] = want
                break
    # la columna amb la season de Zalando porta una data com a capçalera (p.ex. "12/08")
    for c in df.columns:
        if c not in ren and re.fullmatch(r"\d{1,2}/\d{1,2}", str(c).strip()):
            ren[c] = "SEASON ZLD"
    df = df.rename(columns=ren)
    for c in ["SEASON ZLD", "ES POT ENVIAR?"]:
        if c not in df.columns:
            df[c] = np.nan
    df = df[df["model_color"].notna()].copy()
    df["EAN"] = df["EAN"].map(clean_ean)
    df["talla"] = df["talla"].map(lambda t: re.sub(r"\.0$", "", str(t).strip()))
    df["GÈNERE"] = df["GÈNERE"].astype(str).str.strip().str.upper()
    ep = df["ES POT ENVIAR?"].fillna("").astype(str).str.strip()
    df["ES POT ENVIAR?"] = ep.where(~ep.isin(["#N/A", "nan", "None", "<NA>"]), "")
    return df


def load_levels(path: str) -> dict:
    """Retorna {grup: {'levels': [int...], 'table': {level: {talla_str: qty}}}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    groups: dict = {}
    current = None
    levels: list[int] = []
    for r in rows:
        first = r[0]
        if isinstance(first, str) and first.strip() and norm(first) != "nivel":
            current = first.strip().upper()
            groups[current] = {"levels": [], "table": {}}
            levels = []
            continue
        if isinstance(first, str) and norm(first) == "nivel":
            levels = [int(v) for v in r[1:] if isinstance(v, (int, float))]
            groups[current]["levels"] = levels
            for lv in levels:
                groups[current]["table"][lv] = {}
            continue
        if current and levels and isinstance(first, (int, float)):
            talla = str(int(first))
            for lv, v in zip(levels, r[1 : 1 + len(levels)]):
                groups[current]["table"][lv][talla] = int(v) if isinstance(v, (int, float)) else 0
    # normalitza noms de grup (NIÑO pot venir com NINO)
    out = {}
    for g, d in groups.items():
        key = {"MUJER": "MUJER", "CABALLERO": "CABALLERO", "NINO": "NIÑO", "NIÑO": "NIÑO"}.get(norm(g).upper(), g)
        d["totals"] = {lv: int(sum(t.values())) for lv, t in d["table"].items()}  # parells de tot el nivell
        out[key] = d
    return out


def sales_files(data_dir: str) -> list[str]:
    files = sorted(glob.glob(os.path.join(data_dir, "Vendes 20*", "*", "*.xlsx")))
    out = []
    for f in files:
        base = os.path.basename(f)
        if base.startswith("~$"):
            continue
        if any(x in norm(base) for x in EXCLUDE_SALES):
            continue
        if not WEEK_RE.search(base):
            continue
        out.append(f)
    return out


def week_range(path: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    m = WEEK_RE.search(os.path.basename(path))
    d1, m1, d2, m2 = map(int, m.groups())
    ym = re.search(r"Vendes (20\d\d)", path)
    base_year = int(ym.group(1)) if ym else dt.date.today().year
    y1 = base_year - 1 if m1 > m2 else base_year
    return pd.Timestamp(y1, m1, d1), pd.Timestamp(base_year, m2, d2)


def read_dades2(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "DADES2" if "DADES2" in wb.sheetnames else None
    if sheet is None:
        raise ValueError(f"{os.path.basename(path)}: no té pestanya DADES2")
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = [norm(h) if h is not None else "" for h in next(it)]
    idx = {h: i for i, h in enumerate(hdr) if h}
    need = ["model_color", "talla", "initial+shipped"]
    for n in need:
        if n not in idx:
            raise ValueError(f"{os.path.basename(path)}: DADES2 sense columna {n}")
    cm, ct, cq = idx["model_color"], idx["talla"], idx["initial+shipped"]
    cmod, ccol = idx.get("model"), idx.get("color")
    ceur = next((idx[h] for h in idx if h.startswith("venda")), None)
    cret = idx.get("quantity_returned")
    recs = []
    for r in it:
        if r[cm] is None:
            continue
        q = r[cq] if isinstance(r[cq], (int, float)) else 0
        eur = r[ceur] if (ceur is not None and isinstance(r[ceur], (int, float))) else 0.0
        ret = r[cret] if (cret is not None and isinstance(r[cret], (int, float))) else 0
        recs.append((r[cmod] if cmod is not None else None, r[ccol] if ccol is not None else None,
                     str(r[cm]).strip(), str(r[ct]).strip() if r[ct] is not None else "", q, eur, ret))
    df = pd.DataFrame(recs, columns=["MODEL", "COLOR", "MODEL_COLOR", "TALLA", "units", "eur", "returned"])
    df["TALLA"] = df["TALLA"].str.replace(r"\.0$", "", regex=True)
    df["has_eur"] = ceur is not None
    return df


def load_sales(data_dir: str, cache_dir: str) -> tuple[pd.DataFrame, list[dict]]:
    os.makedirs(cache_dir, exist_ok=True)
    frames, sources = [], []
    for f in sales_files(data_dir):
        start, end = week_range(f)
        mtime = int(os.path.getmtime(f))
        cache = os.path.join(cache_dir, f"{os.path.basename(f)}.{mtime}.parquet")
        if os.path.exists(cache):
            df = pd.read_parquet(cache)
        else:
            df = read_dades2(f)
            try:
                df.to_parquet(cache)
            except Exception:
                pass
        df = df.assign(start=start, end=end, file=os.path.basename(f))
        frames.append(df)
        sources.append({"fitxer": os.path.basename(f), "inici": start.date().isoformat(), "fi": end.date().isoformat(),
                        "unitats": int(df["units"].sum()), "eur": round(float(df["eur"].sum()), 2),
                        "amb_eur": bool(df["has_eur"].iloc[0]) if len(df) else False})
    if not frames:
        raise SystemExit("No s'ha trobat cap fitxer de vendes setmanals a 'Vendes 20xx/'")
    lines = pd.concat(frames, ignore_index=True)
    # setmanes duplicades (mateix rang en dos fitxers) -> avisa i queda't amb el primer
    dup = lines.drop_duplicates(["file", "start", "end"]).duplicated(["start", "end"], keep="first")
    if dup.any():
        bad = lines.drop_duplicates(["file", "start", "end"])[dup]["file"].tolist()
        print("AVÍS: setmanes duplicades, s'ignoren:", bad)
        lines = lines[~lines["file"].isin(bad)]
    return lines, sorted(sources, key=lambda s: s["inici"])


def load_sales_2025(path: str) -> pd.Series:
    raw = pd.read_excel(path, header=None)
    hdr_row = None
    for i in range(min(10, len(raw))):
        if any(norm(v) == "etiquetas de fila" for v in raw.iloc[i].tolist() if isinstance(v, str)):
            hdr_row = i
            break
    if hdr_row is None:
        hdr_row = 0
    df = raw.iloc[hdr_row + 1 :, :2].copy()
    df.columns = ["model_color", "units"]
    df = df[df["model_color"].notna() & (df["model_color"].astype(str) != "Total general")]
    df["units"] = pd.to_numeric(df["units"], errors="coerce").fillna(0)
    return df.groupby(df["model_color"].astype(str).str.strip())["units"].sum()


def load_stock_tp(folder: str) -> tuple[pd.DataFrame, str]:
    files = sorted(glob.glob(os.path.join(folder, "*.txt")), key=os.path.getmtime)
    if not files:
        raise SystemExit(f"No hi ha cap .txt a {folder}")
    path = files[-1]
    df = pd.read_csv(path, sep="\t", encoding="utf-16", dtype=str)
    cols = {norm(c): c for c in df.columns}
    ean_col = cols.get("codigo de barras") or cols.get("código de barras") or cols.get("ean")
    c0102 = cols.get("stock 01 02")
    c30 = cols.get("stock disponible 30 dies")
    c59 = cols.get("stock disponible 59 dies")
    if not all([ean_col, c0102, c30, c59]):
        raise SystemExit(f"Columnes no trobades a {os.path.basename(path)}: {list(df.columns)}")
    out = pd.DataFrame({
        "EAN": df[ean_col].map(clean_ean),
        "STOCK TP 01 02": to_num(df[c0102]),
        "DISPO 30 DIES": to_num(df[c30]),
        "DISPO 59 DIES": to_num(df[c59]),
    })
    out = out[out["EAN"].notna()].groupby("EAN", as_index=False).sum()
    return out, os.path.basename(path)


def _snapshot_frame(path: str) -> pd.DataFrame | None:
    """Llegeix un snapshot (csv o xlsx) i retorna EAN / Offerable / Non-offerable / Total, o None si no és vàlid."""
    ext = os.path.splitext(path)[1].lower()
    df = None
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", errors="replace") as fh:
            head = fh.readline()
        sep = ";" if head.count(";") >= head.count(",") else ","
        df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig", encoding_errors="replace")
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            first = next(it, None)
            if first and any(norm(v) == "ean" for v in first if v is not None):
                hdr = [str(v) if v is not None else f"c{i}" for i, v in enumerate(first)]
                df = pd.DataFrame(list(it), columns=hdr)
                break
    if df is None or df.empty:
        return None
    cols = {norm(c): c for c in df.columns}
    ean_col = cols.get("ean")
    if ean_col is None:
        return None
    off = cols.get("offerable stock")
    non = cols.get("non-offerable stock")
    tot = cols.get("total")
    if off is None or tot is None:
        return None
    out = pd.DataFrame({
        "EAN": df[ean_col].map(clean_ean),
        "OFFERABLE": to_num(df[off]),
        "NON OFFERABLE": to_num(df[non]) if non else 0,
        "STOCK ZLD": to_num(df[tot]),
    })
    valid = out["EAN"].notna().mean()
    if valid < 0.9:   # EANs en notació científica (8,43453E+12) o buits
        return None
    return out[out["EAN"].notna()].groupby("EAN", as_index=False).sum()


def load_snapshot(folder: str, extra_paths: list[str] | None = None) -> tuple[pd.DataFrame, dict]:
    cands = sorted(glob.glob(os.path.join(folder, "*.csv")) + glob.glob(os.path.join(folder, "*.xlsx")))
    cands = [c for c in cands if not os.path.basename(c).startswith("~$")]
    if extra_paths:
        cands += [p for p in extra_paths if os.path.exists(p)]
    dated = []
    for c in cands:
        m = SNAP_DATE_RE.search(os.path.basename(c))
        d = dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else dt.date.fromtimestamp(os.path.getmtime(c))
        dated.append((d, c))
    rejected, chosen, frame = [], None, None
    for d, c in sorted(dated, reverse=True):
        fr = _snapshot_frame(c)
        if fr is None:
            rejected.append(os.path.basename(c))
            continue
        chosen, frame = (d, c), fr
        break
    if frame is None:
        raise SystemExit("Cap snapshot de stock Zalando vàlid (EANs íntegres). Rebutjats: " + ", ".join(rejected))
    meta = {"fitxer": os.path.basename(chosen[1]), "data": chosen[0].isoformat(), "rebutjats": rejected,
            "eans": int(len(frame)), "total": int(frame["STOCK ZLD"].sum())}
    return frame, meta


def load_pending(folder: str) -> tuple[pd.DataFrame, list[str]]:
    files = sorted(glob.glob(os.path.join(folder, "*.csv")))
    frames, labels = [], []
    for f in files:
        base = os.path.basename(f)
        with open(f, "r", encoding="utf-8-sig", errors="replace") as fh:
            head = fh.readline()
        sep = ";" if head.count(";") >= head.count(",") else ","
        df = pd.read_csv(f, sep=sep, dtype=str, encoding="utf-8-sig", encoding_errors="replace")
        cols = {norm(c): c for c in df.columns}
        ec, qc = cols.get("ean"), cols.get("quantity") or cols.get("quantitat") or cols.get("qty")
        if ec is None or qc is None:
            print(f"AVÍS: {base} sense columnes ean/quantity, s'ignora")
            continue
        m = re.search(r"(\d\d)(\d\d)(\d{4})", base)
        label = f"ENV {m.group(1)}.{m.group(2)}" if m else f"ENV {os.path.splitext(base)[0][:12]}"
        while label in labels:
            label += "'"
        labels.append(label)
        g = pd.DataFrame({"EAN": df[ec].map(clean_ean), label: to_num(df[qc])}).dropna(subset=["EAN"]).groupby("EAN", as_index=False).sum()
        frames.append(g)
    if not frames:
        return pd.DataFrame({"EAN": pd.Series(dtype=str), "ENV PENDENTS": pd.Series(dtype=float)}), []
    out = frames[0]
    for fr in frames[1:]:
        out = out.merge(fr, on="EAN", how="outer")
    out = out.fillna(0)
    out["ENV PENDENTS"] = out[labels].sum(axis=1)
    return out, labels


def load_adjustments(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        return pd.DataFrame(columns=["model_color", "multiplicador", "nivell", "comentari"])
    df = pd.read_excel(path)
    cols = {norm(c): c for c in df.columns}
    out = pd.DataFrame({
        "model_color": df[cols["model_color"]].fillna("").astype(str).str.strip(),
        "multiplicador": pd.to_numeric(df[cols["multiplicador"]], errors="coerce") if "multiplicador" in cols else np.nan,
        "nivell": pd.to_numeric(df[cols["nivell"]], errors="coerce") if "nivell" in cols else np.nan,
        "comentari": df[cols["comentari"]] if "comentari" in cols else "",
    })
    return out[out["model_color"].ne("")]


# --------------------------------------------------------------------------------------
# Càlcul
# --------------------------------------------------------------------------------------
def pick_level(levels: list[int], totals: dict, target: float, min_level: int) -> tuple[int | None, str]:
    """Primer nivell (ordenats pel total de parells) el total del qual cobreix l'objectiu."""
    order = sorted(levels, key=lambda l: (totals.get(l, 0), l))
    if target <= 0:
        if min_level:
            lv = min_level if min_level in levels else next((l for l in order if l >= min_level), None)
            return lv, "nivell mínim"
        return None, ""
    for lv in order:
        if totals.get(lv, 0) >= target:
            if min_level and min_level in levels and totals.get(min_level, 0) > totals.get(lv, 0):
                return min_level, "nivell mínim"
            return lv, ""
    top = order[-1]
    return top, f"objectiu {int(target)} > màxim de la taula ({totals.get(top, 0)} parells, nivell {top})"


def size_qty(table: dict, group: str, gender: str, talla: str) -> tuple[int, str]:
    if talla in table:
        return table[talla], ""
    if gender == "UNISEX" and talla.isdigit() and int(talla) > 45 and "45" in table:
        return table["45"], "talles 46-47 = talla 45"
    if group == "NIÑO" and talla.isdigit() and int(talla) < 25 and "25" in table:
        return table["25"], "talles <25 = talla 25"
    return 0, f"talla {talla} fora de taula"


def compute(models: pd.DataFrame, levels: dict, lines: pd.DataFrame, acum25: pd.Series,
            stock_tp: pd.DataFrame, snap: pd.DataFrame, pending: pd.DataFrame, pend_labels: list[str],
            adjust: pd.DataFrame, mult: float, min_level: int, min_level_kids: int, max_level: int | None = None):
    last_end = lines["end"].max()
    last_start = lines.loc[lines["end"] == last_end, "start"].iloc[0]
    week_ends = sorted(lines["end"].unique())
    last4 = week_ends[-4:]
    lw = lines[lines["end"] == last_end]
    prev_end = week_ends[-2] if len(week_ends) > 1 else None
    pw = lines[lines["end"] == prev_end] if prev_end is not None else lines.iloc[0:0]

    mc_week = lw.groupby("MODEL_COLOR")["units"].sum()
    mc_prev = pw.groupby("MODEL_COLOR")["units"].sum()
    mc_4w = lines[lines["end"].isin(last4)].groupby("MODEL_COLOR")["units"].sum()
    mc_26 = lines.groupby("MODEL_COLOR")["units"].sum()
    sku_week = lw.groupby(["MODEL_COLOR", "TALLA"])["units"].sum()
    sku_26 = lines.groupby(["MODEL_COLOR", "TALLA"])["units"].sum()

    adj = adjust.set_index("model_color") if len(adjust) else adjust

    df = models.copy()
    df["VENDA 1 SETM"] = df["model_color"].map(mc_week).fillna(0).astype(int)
    df["VENDA SETM ANT"] = df["model_color"].map(mc_prev).fillna(0).astype(int)
    df["VENDA 4 SETM"] = df["model_color"].map(mc_4w).fillna(0).astype(int)
    df["ACUM'25"] = df["model_color"].map(acum25).fillna(0).astype(int)
    df["ACUM'26"] = df["model_color"].map(mc_26).fillna(0).astype(int)
    key = list(zip(df["model_color"], df["talla"]))
    df["VENDA SKU 1 SETM"] = [int(sku_week.get(k, 0)) for k in key]
    df["VENDA SKU ACUM'26"] = [int(sku_26.get(k, 0)) for k in key]

    # multiplicador i nivell per model_color
    mults, forced = {}, {}
    for mc in df["model_color"].unique():
        m = mult
        f = None
        if len(adj) and mc in adj.index:
            row = adj.loc[mc]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            if pd.notna(row.get("multiplicador", np.nan)):
                m = float(row["multiplicador"])
            if pd.notna(row.get("nivell", np.nan)):
                f = int(row["nivell"])
        mults[mc], forced[mc] = m, f
    df["MULT"] = df["model_color"].map(mults)
    df["OBJECTIU"] = (df["VENDA 1 SETM"] * df["MULT"]).round().astype(int)

    grp = [GENDER_GROUP.get(g, "?") for g in df["GÈNERE"]]
    df["GRUP NIVELL"] = [g or "" for g in grp]
    niv, hauria, avis = [], [], []
    mc_cache: dict = {}
    for i, r in enumerate(df.itertuples(index=False)):
        mc, gender, g, talla, target = r.model_color, r.GÈNERE, grp[i], r.talla, r.OBJECTIU
        notes = []
        if g == "?":
            niv.append(None); hauria.append(0); avis.append(f"gènere {gender} sense taula assignada")
            continue
        if g is None:
            # sense taula (complements): objectiu directe si talla única
            q = int(target) if talla.upper() in ("UNICA", "ÚNICA", "UNICO", "U") else 0
            niv.append(None); hauria.append(q); avis.append("sense taula de nivells")
            continue
        if g not in levels:
            niv.append(None); hauria.append(0); avis.append(f"grup {g} no és a NIVEL.xlsx")
            continue
        if mc not in mc_cache:
            lvls = [l for l in levels[g]["levels"] if max_level is None or l <= max_level] or levels[g]["levels"][:1]
            ml = min_level_kids if g in KIDS_GROUPS else min_level
            if forced.get(mc) is not None:
                lv = forced[mc] if forced[mc] in lvls else next((l for l in lvls if l >= forced[mc]), lvls[-1])
                mc_cache[mc] = (lv, f"nivell forçat {forced[mc]}")
            else:
                mc_cache[mc] = pick_level(lvls, levels[g]["totals"], target, ml)
        lv, note = mc_cache[mc]
        if note:
            notes.append(note)
        if lv is None:
            niv.append(None); hauria.append(0); avis.append("; ".join(notes))
            continue
        q, n2 = size_qty(levels[g]["table"][lv], g, gender, talla)
        if n2:
            notes.append(n2)
        niv.append(lv); hauria.append(int(q)); avis.append("; ".join(notes))
    df["NIVELL"] = pd.array(niv, dtype="Int64")
    df["PARELLS NIVELL"] = pd.array([levels[g]["totals"].get(lv) if (lv is not None and g in levels) else None
                                     for g, lv in zip(grp, niv)], dtype="Int64")
    df["HAURIA"] = hauria
    df["AVÍS"] = avis

    # stocks
    df = df.merge(snap, on="EAN", how="left").merge(pending, on="EAN", how="left").merge(stock_tp, on="EAN", how="left")
    for c in ["STOCK ZLD", "OFFERABLE", "NON OFFERABLE", "ENV PENDENTS", "STOCK TP 01 02", "DISPO 30 DIES", "DISPO 59 DIES"] + pend_labels:
        if c not in df.columns:
            df[c] = 0
        df[c] = df[c].fillna(0).astype(int)
    df.loc[df["EAN"].isna(), "AVÍS"] = (df.loc[df["EAN"].isna(), "AVÍS"] + "; sense EAN").str.strip("; ")

    # enviable? (HI26 NOU sense marca a 'es pot enviar?' = encara no creat a Zalando)
    not_created = (df["SEASON"].astype(str).str.upper() == "HI26") & (df["TEMPORADA"].astype(str).str.upper() == "NOU") & (df["ES POT ENVIAR?"] == "")
    df["CREAT A ZLD?"] = np.where(not_created, "NO CONSTA", "SÍ")

    df["DIF"] = df["HAURIA"] - df["STOCK ZLD"] - df["ENV PENDENTS"]
    df["REPO"] = np.where(df["CREAT A ZLD?"] == "SÍ", df["DIF"].clip(lower=0), 0).astype(int)
    df["PREPARABLE"] = np.minimum(df["REPO"], df["DISPO 59 DIES"]).astype(int)
    df["FALTA STOCK TP"] = (df["REPO"] - df["PREPARABLE"]).astype(int)

    def talla_key(t):
        return (0, int(t)) if str(t).isdigit() else (1, str(t))
    df["_tk"] = df["talla"].map(talla_key)
    df = df.sort_values(["model_color", "_tk"]).drop(columns="_tk").reset_index(drop=True)

    sku_cols = ["EAN", "SKU", "SEASON", "TEMPORADA", "COL·LECCIÓ", "GÈNERE", "model", "color", "model_color", "talla",
                "SEASON ZLD", "ES POT ENVIAR?", "CREAT A ZLD?", "VENDA 1 SETM", "ACUM'25", "ACUM'26", "VENDA 4 SETM",
                "MULT", "OBJECTIU", "NIVELL", "HAURIA", "STOCK ZLD", "OFFERABLE", "NON OFFERABLE"] + pend_labels + \
               ["ENV PENDENTS", "DIF", "REPO", "STOCK TP 01 02", "DISPO 30 DIES", "PREPARABLE",
                "VENDA SKU 1 SETM", "VENDA SKU ACUM'26", "AVÍS"]
    sku = df[sku_cols].copy()

    # vista model_color
    first = {c: "first" for c in ["model", "color", "SEASON", "TEMPORADA", "COL·LECCIÓ", "GÈNERE", "SEASON ZLD", "ES POT ENVIAR?",
                                 "CREAT A ZLD?", "VENDA 1 SETM", "VENDA SETM ANT", "VENDA 4 SETM", "MULT", "OBJECTIU", "GRUP NIVELL",
                                 "NIVELL", "PARELLS NIVELL", "ACUM'25", "ACUM'26"]}
    sums = {c: "sum" for c in ["HAURIA", "STOCK ZLD", "OFFERABLE", "ENV PENDENTS", "DIF", "REPO", "PREPARABLE", "FALTA STOCK TP",
                               "STOCK TP 01 02", "DISPO 30 DIES", "DISPO 59 DIES"]}
    mc = df.groupby("model_color").agg({**first, **sums, "talla": "count"}).rename(columns={"talla": "N TALLES"})
    mc["TALLES AMB REPO"] = df[df["REPO"] > 0].groupby("model_color").size().reindex(mc.index).fillna(0).astype(int)
    mc["TALLES SENSE STOCK ZLD"] = df[(df["STOCK ZLD"] + df["ENV PENDENTS"]) <= 0].groupby("model_color").size().reindex(mc.index).fillna(0).astype(int)
    cov = (mc["STOCK ZLD"] + mc["ENV PENDENTS"]) / mc["VENDA 1 SETM"].replace(0, np.nan)
    mc["COBERTURA SETM"] = cov.round(1)
    mc["AVÍS"] = df.groupby("model_color")["AVÍS"].agg(lambda s: "; ".join(sorted({x for x in s if x})))
    mc = mc.reset_index()
    mc_cols = ["model_color", "model", "color", "SEASON", "TEMPORADA", "COL·LECCIÓ", "GÈNERE", "SEASON ZLD", "CREAT A ZLD?",
               "VENDA 1 SETM", "ACUM'25", "ACUM'26", "VENDA 4 SETM", "MULT", "OBJECTIU", "NIVELL", "PARELLS NIVELL", "HAURIA",
               "STOCK ZLD", "OFFERABLE", "ENV PENDENTS", "COBERTURA SETM", "DIF", "REPO", "PREPARABLE",
               "STOCK TP 01 02", "DISPO 30 DIES", "AVÍS"]
    mc = mc[mc_cols].sort_values(["REPO", "VENDA 1 SETM", "ACUM'26"], ascending=[False, False, False]).reset_index(drop=True)

    # vendes de la setmana de model_colors fora de la llista
    fora = lw[~lw["MODEL_COLOR"].isin(set(models["model_color"]))].groupby("MODEL_COLOR").agg(
        MODEL=("MODEL", "first"), COLOR=("COLOR", "first"), **{"VENDA 1 SETM": ("units", "sum")}).reset_index()
    fora["ACUM'26"] = fora["MODEL_COLOR"].map(mc_26).fillna(0).astype(int)
    fora["ACUM'25"] = fora["MODEL_COLOR"].map(acum25).fillna(0).astype(int)
    fora = fora.sort_values("VENDA 1 SETM", ascending=False).reset_index(drop=True)

    info = {"setmana_inici": last_start.date().isoformat(), "setmana_fi": last_end.date().isoformat(),
            "setmanes": len(week_ends), "setmana_ant": (prev_end.date().isoformat() if prev_end is not None else "")}
    return sku, mc, fora, info


# --------------------------------------------------------------------------------------
# Sortides
# --------------------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")
REPO_FILL = PatternFill("solid", fgColor="E2F0D9")
NO_FILL = PatternFill("solid", fgColor="EDEDED")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
GREY_HDR = "D9D9D9"
YELLOW_HDR = "FFE699"


def style_sheet(ws, df: pd.DataFrame, highlight_col: str | None = None, grey_col: str | None = None, key_cols=(), header_groups=()):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    names = list(df.columns)
    hdr_fill = {}
    for start, end, color in header_groups:
        if start in names and end in names:
            for k in range(names.index(start), names.index(end) + 1):
                hdr_fill[k + 1] = PatternFill("solid", fgColor=color)
    for j, c in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j)
        cell.fill = hdr_fill.get(j, HDR_FILL)
        cell.font = Font(bold=True, color="000000") if j in hdr_fill else HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if c in key_cols:
            for i in range(2, len(df) + 2):
                ws.cell(row=i, column=j).fill = KEY_FILL
        width = max(8, min(38, int(max([len(str(c))] + [len(str(v)) for v in df[c].head(300).tolist()]) * 1.1) + 1))
        if c in ("AVÍS",):
            width = 40
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.row_dimensions[1].height = 32
    if highlight_col and highlight_col in df.columns:
        j = list(df.columns).index(highlight_col) + 1
        for i, v in enumerate(df[highlight_col].tolist(), start=2):
            if isinstance(v, (int, float)) and v > 0:
                ws.cell(row=i, column=j).fill = REPO_FILL
                ws.cell(row=i, column=j).font = Font(bold=True)
    if grey_col and grey_col in df.columns:
        j = list(df.columns).index(grey_col) + 1
        for i, v in enumerate(df[grey_col].tolist(), start=2):
            if v == "NO CONSTA":
                ws.cell(row=i, column=j).fill = NO_FILL


def write_vendes(lines: pd.DataFrame, sources: list[dict], acum25: pd.Series, path: str, info: dict):
    last_end = lines["end"].max()
    lw = lines[lines["end"] == last_end]
    mc = lines.groupby("MODEL_COLOR").agg(MOD=("MODEL", "first"), COL=("COLOR", "first"), **{
        "2026 ACUMULAT": ("units", "sum"), "2026 EUR (des de 18.03)": ("eur", "sum"), "RETORNS 2026": ("returned", "sum")}).reset_index()
    mc = mc.rename(columns={"MODEL_COLOR": "model_color"})
    mc["2026 EUR (des de 18.03)"] = mc["2026 EUR (des de 18.03)"].round(2)
    wk_label = f"SETM {info['setmana_inici'][8:10]}.{info['setmana_inici'][5:7]}-{info['setmana_fi'][8:10]}.{info['setmana_fi'][5:7]}"
    mc[wk_label] = mc["model_color"].map(lw.groupby("MODEL_COLOR")["units"].sum()).fillna(0).astype(int)
    mc["2025 ACUMULAT"] = mc["model_color"].map(acum25).fillna(0).astype(int)
    mc = mc.sort_values("2026 ACUMULAT", ascending=False).reset_index(drop=True)
    sku = lines.groupby(["MODEL_COLOR", "TALLA"]).agg(**{"2026 ACUMULAT": ("units", "sum")}).reset_index()
    lw_sku = lw.groupby(["MODEL_COLOR", "TALLA"])["units"].sum()
    sku[wk_label] = [int(lw_sku.get(k, 0)) for k in zip(sku["MODEL_COLOR"], sku["TALLA"])]
    sku = sku.sort_values(["2026 ACUMULAT"], ascending=False).reset_index(drop=True)
    piv = lines.pivot_table(index="MODEL_COLOR", columns="end", values="units", aggfunc="sum", fill_value=0)
    piv.columns = [f"{c.day:02d}.{c.month:02d}" for c in piv.columns]
    piv["TOTAL"] = piv.sum(axis=1)
    piv = piv.sort_values("TOTAL", ascending=False).reset_index().rename(columns={"MODEL_COLOR": "model_color"})
    src = pd.DataFrame(sources).rename(columns={"fitxer": "FITXER", "inici": "INICI", "fi": "FI", "unitats": "UNITATS", "eur": "EUR", "amb_eur": "TÉ VENDA €"})
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        mc.to_excel(xw, sheet_name="MODEL_COLOR", index=False)
        sku.to_excel(xw, sheet_name="SKU", index=False)
        piv.to_excel(xw, sheet_name="SETMANES", index=False)
        src.to_excel(xw, sheet_name="FONTS", index=False)
        for name, d in [("MODEL_COLOR", mc), ("SKU", sku), ("SETMANES", piv), ("FONTS", src)]:
            style_sheet(xw.sheets[name], d)
    return mc


def write_excel(sku: pd.DataFrame, mc: pd.DataFrame, fora: pd.DataFrame, params: list[tuple], levels: dict, path: str):
    lvl_rows = []
    for g, d in levels.items():
        lvl_rows.append([g] + d["levels"])
        sizes = sorted({t for lv in d["table"].values() for t in lv}, key=lambda t: (len(t), t))
        for t in sizes:
            lvl_rows.append([t] + [d["table"][lv].get(t, 0) for lv in d["levels"]])
        lvl_rows.append(["TOTAL PARELLS"] + [d["totals"][lv] for lv in d["levels"]])
        lvl_rows.append([])
    width = max(len(r) for r in lvl_rows)
    lvl = pd.DataFrame([r + [None] * (width - len(r)) for r in lvl_rows], columns=["NIVELL"] + [f"c{i}" for i in range(1, width)])
    par = pd.DataFrame(params, columns=["PARÀMETRE", "VALOR"])
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        sku.to_excel(xw, sheet_name="CÀLCUL SKU", index=False)
        mc.to_excel(xw, sheet_name="MODEL_COLOR", index=False)
        fora.to_excel(xw, sheet_name="FORA LLISTA", index=False)
        par.to_excel(xw, sheet_name="PARÀMETRES", index=False)
        lvl.to_excel(xw, sheet_name="NIVELLS", index=False, header=False)
        groups_sku = (("EAN", "CREAT A ZLD?", GREY_HDR), ("VENDA 1 SETM", "OBJECTIU", YELLOW_HDR))
        groups_mc = (("model_color", "CREAT A ZLD?", GREY_HDR), ("VENDA 1 SETM", "OBJECTIU", YELLOW_HDR))
        style_sheet(xw.sheets["CÀLCUL SKU"], sku, highlight_col="REPO", grey_col="CREAT A ZLD?", key_cols=("HAURIA", "DIF", "REPO", "PREPARABLE"), header_groups=groups_sku)
        style_sheet(xw.sheets["MODEL_COLOR"], mc, highlight_col="REPO", grey_col="CREAT A ZLD?", key_cols=("HAURIA", "REPO", "PREPARABLE"), header_groups=groups_mc)
        style_sheet(xw.sheets["FORA LLISTA"], fora)
        style_sheet(xw.sheets["PARÀMETRES"], par)
        xw.sheets["PARÀMETRES"].column_dimensions["A"].width = 34
        xw.sheets["PARÀMETRES"].column_dimensions["B"].width = 110


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="ca"><head><meta charset="utf-8"><title>__TITLE__</title>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
:root{--bg:#f6f7f9;--card:#fff;--ink:#1c2430;--muted:#66717f;--line:#e3e7ec;--head:#1f3864;--accent:#2e7d32;--warn:#b26a00;--bad:#b3261e}
*{box-sizing:border-box}body{margin:0;font:14px/1.4 system-ui,Segoe UI,Roboto,Arial,sans-serif;color:var(--ink);background:var(--bg)}
header{background:var(--head);color:#fff;padding:14px 22px}header h1{margin:0;font-size:19px;font-weight:600}
header .sub{opacity:.85;font-size:12.5px;margin-top:4px}
.warn{background:#fff4e5;border-left:4px solid var(--warn);color:#5a3a00;padding:8px 14px;margin:12px 22px 0;border-radius:4px;font-size:13px}
.tabs{display:flex;gap:6px;padding:14px 22px 0}
.tab{border:1px solid var(--line);border-bottom:none;background:#e9edf2;padding:8px 16px;border-radius:8px 8px 0 0;cursor:pointer;font-weight:600;color:var(--muted)}
.tab.active{background:var(--card);color:var(--ink)}
.panel{display:none;background:var(--card);margin:0 22px 22px;border:1px solid var(--line);border-radius:0 8px 8px 8px;padding:12px}
.panel.active{display:block}
.bar{display:flex;flex-wrap:wrap;gap:10px;align-items:center;margin-bottom:10px}
.bar input[type=text]{padding:7px 10px;border:1px solid var(--line);border-radius:6px;min-width:260px}
.bar select{padding:6px 8px;border:1px solid var(--line);border-radius:6px;background:#fff}
.bar label{font-size:13px;color:var(--muted);display:flex;gap:4px;align-items:center}
.kpis{display:flex;flex-wrap:wrap;gap:10px;margin:0 0 10px}
.kpi{background:#f1f4f8;border-radius:8px;padding:8px 14px;min-width:120px}.kpi b{display:block;font-size:20px}.kpi span{font-size:12px;color:var(--muted)}
.wrap{overflow:auto;max-height:78vh;border:1px solid var(--line);border-radius:0 0 6px 6px}
table{border-collapse:separate;border-spacing:0;width:max-content;min-width:100%;font-size:12.5px}
th{position:sticky;top:0;background:var(--head);color:#fff;padding:6px 8px;text-align:left;cursor:pointer;white-space:nowrap;user-select:none;z-index:2;border-right:1px solid rgba(255,255,255,.12)}
th.num{text-align:right}th .arr{opacity:.7;font-size:10px;margin-left:3px}
th.hg-grey{background:#d9d9d9;color:#1c2430}th.hg-yellow{background:#ffe699;color:#1c2430}
[hidden]{display:none!important}
.btn{padding:7px 12px;border:1px solid var(--line);border-radius:6px;background:#fff;cursor:pointer;font-weight:600;color:var(--ink)}
.colwrap{position:relative}
.colpick{position:absolute;top:110%;left:0;z-index:20;background:#fff;border:1px solid var(--line);border-radius:8px;box-shadow:0 8px 24px rgba(0,0,0,.15);padding:10px;min-width:560px;max-height:60vh;overflow:auto}
.cp-grid{display:grid;grid-template-columns:repeat(3,minmax(160px,1fr));gap:4px 14px;font-size:12.5px}
.cp-grid label{display:flex;gap:6px;align-items:center;white-space:nowrap;cursor:pointer}
.cp-actions{display:flex;gap:8px;align-items:center;margin-bottom:8px;font-size:12px;color:var(--muted)}
.cp-actions button{padding:4px 10px;border:1px solid var(--line);border-radius:6px;background:#f1f4f8;cursor:pointer}
.topscroll{overflow-x:auto;overflow-y:hidden;height:18px;border:1px solid var(--line);border-bottom:none;border-radius:6px 6px 0 0;background:#fff}
.topscroll>div{height:1px}
.wrap::-webkit-scrollbar,.topscroll::-webkit-scrollbar{height:14px;width:14px}
.wrap::-webkit-scrollbar-thumb,.topscroll::-webkit-scrollbar-thumb{background:#8a97a6;border-radius:8px;border:3px solid #fff}
.wrap::-webkit-scrollbar-track,.topscroll::-webkit-scrollbar-track{background:#e9edf2}
th.sticky{left:0;z-index:4}
td.sticky{position:sticky;left:0;background:#fff;z-index:1;box-shadow:1px 0 0 var(--line)}
tr:hover td.sticky{background:#f4f7fb}
tfoot td.sticky{background:#eef2f6;z-index:3}
td{padding:4px 8px;border-bottom:1px solid var(--line);white-space:nowrap}td.num{text-align:right;font-variant-numeric:tabular-nums}
tr:hover td{background:#f4f7fb}td.repo{background:#e2f0d9;font-weight:600}td.no{color:var(--bad)}td.neg{color:#8a94a0}
tfoot td{position:sticky;bottom:0;background:#eef2f6;font-weight:600;border-top:2px solid #c9d1db;z-index:1}
td.key{background:#fffbea}
.muted{color:var(--muted);font-size:12px}
.count{margin-left:auto;font-size:12.5px;color:var(--muted)}
</style></head><body>
<header><h1>__TITLE__</h1><div class="sub">__SUBTITLE__</div></header>
__WARNINGS__
<div class="tabs"><div class="tab active" data-t="mc">Per model_color</div><div class="tab" data-t="sku">Detall per SKU</div></div>
<div id="p-mc" class="panel active"></div>
<div id="p-sku" class="panel"></div>
<script>
const DATA = __DATA__;
const NUMFMT = new Intl.NumberFormat('ca-ES');
function esc(v){ return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/"/g,'&quot;'); }
function build(id, spec, rows){
  const panel = document.getElementById('p-'+id);
  const storeKey = 'repo-zld-cols-'+id;
  let hidden = new Set();
  try { const s = localStorage.getItem(storeKey); if(s) hidden = new Set(JSON.parse(s)); } catch(e) {}
  const state = {q:'', sortKey: spec.defaultSort, sortDir: -1, onlyRepo: spec.onlyRepoDefault, filters:{}};
  const facets = spec.facets.map(f => ({key:f, values:[...new Set(rows.map(r=>r[f]).filter(v=>v!==null && v!==''))].sort()}));
  let html = '<div class="kpis" id="k-'+id+'"></div><div class="bar">';
  html += '<input type="text" id="q-'+id+'" placeholder="Cerca (model, color, EAN, SKU...)">';
  facets.forEach(f => { html += '<select data-f="'+esc(f.key)+'"><option value="">'+esc(f.key)+': tots</option>'+f.values.map(v=>'<option>'+esc(v)+'</option>').join('')+'</select>'; });
  html += '<label><input type="checkbox" id="r-'+id+'" '+(state.onlyRepo?'checked':'')+'> només REPO &gt; 0</label>';
  html += '<div class="colwrap"><button type="button" class="btn" id="cb-'+id+'">Columnes ▾</button><div class="colpick" id="cp-'+id+'" hidden></div></div>';
  html += '<span class="count" id="c-'+id+'"></span></div>';
  html += '<div class="topscroll" id="ts-'+id+'"><div></div></div>';
  html += '<div class="wrap" id="w-'+id+'"><table id="t-'+id+'"><thead><tr></tr></thead><tbody></tbody><tfoot><tr></tr></tfoot></table></div>';
  panel.innerHTML = html;
  const wrap = panel.querySelector('.wrap'), table = panel.querySelector('table');
  const thead = panel.querySelector('thead tr'), tbody = panel.querySelector('tbody'), tfoot = panel.querySelector('tfoot tr');
  const topscroll = panel.querySelector('.topscroll'), topinner = topscroll.firstElementChild;
  const colpick = panel.querySelector('.colpick'), colbtn = panel.querySelector('#cb-'+id);
  function visCols(){ return spec.cols.filter(c => !hidden.has(c.k)); }
  function saveCols(){ try { localStorage.setItem(storeKey, JSON.stringify([...hidden])); } catch(e) {} }
  function setHidden(next){ hidden = next; saveCols(); renderPicker(); renderHead(); render(); }
  function renderPicker(){
    let h = '<div class="cp-actions"><button type="button" data-a="all">Totes</button><button type="button" data-a="none">Cap</button>';
    h += '<span>'+(spec.cols.length-hidden.size)+' de '+spec.cols.length+' columnes visibles</span></div><div class="cp-grid">';
    spec.cols.forEach(c => { h += '<label><input type="checkbox" data-c="'+esc(c.k)+'" '+(hidden.has(c.k)?'':'checked')+'> '+esc(c.l)+'</label>'; });
    colpick.innerHTML = h + '</div>';
    colpick.querySelectorAll('input').forEach(i => i.addEventListener('change', e => {
      const k = e.target.dataset.c; const next = new Set(hidden);
      if(e.target.checked) next.delete(k); else next.add(k);
      if(next.size >= spec.cols.length){ e.target.checked = true; return; }
      setHidden(next);
    }));
    colpick.querySelectorAll('button').forEach(b => b.addEventListener('click', () => {
      if(b.dataset.a === 'all') setHidden(new Set());
      else setHidden(new Set(spec.cols.slice(1).map(c => c.k)));
    }));
  }
  function renderHead(){
    const cols = visCols();
    thead.innerHTML = cols.map((c,i) => '<th class="'+(c.n?'num':'')+(c.hg?' hg-'+c.hg:'')+(i===0?' sticky':'')+'" data-k="'+esc(c.k)+'" title="'+esc(c.l)+'">'+esc(c.l)+'<span class="arr"></span></th>').join('');
    thead.querySelectorAll('th').forEach(th => th.addEventListener('click', () => {
      const k = th.dataset.k;
      if(state.sortKey===k) state.sortDir*=-1; else { state.sortKey=k; state.sortDir = spec.cols.find(c=>c.k===k).n ? -1 : 1; }
      render();
    }));
  }
  function fitHeight(){
    const top = wrap.getBoundingClientRect().top;
    wrap.style.maxHeight = Math.max(240, window.innerHeight - top - 16) + 'px';
    topinner.style.width = table.scrollWidth + 'px';
  }
  function render(){
    const cols = visCols();
    const q = state.q.toLowerCase();
    let out = rows.filter(r => {
      if(state.onlyRepo && !(r.REPO > 0)) return false;
      for(const k in state.filters){ if(state.filters[k] && String(r[k]) !== state.filters[k]) return false; }
      if(!q) return true;
      return spec.search.some(k => r[k] !== null && String(r[k]).toLowerCase().includes(q));
    });
    const sk = state.sortKey, sd = state.sortDir;
    out.sort((a,b) => { const x=a[sk], y=b[sk]; if(x===y) return 0; if(x===null||x===undefined) return 1; if(y===null||y===undefined) return -1; return (x>y?1:-1)*sd; });
    thead.querySelectorAll('th').forEach(th => th.querySelector('.arr').textContent = th.dataset.k===sk ? (sd>0?'▲':'▼') : '');
    const MAX = 6000; const shown = out.slice(0, MAX);
    let h = '';
    for(const r of shown){
      h += '<tr>';
      cols.forEach((c,i) => {
        let v = r[c.k]; let cls = c.n ? 'num' : '';
        if(c.k === 'REPO' && v > 0) cls += ' repo';
        if(c.k === 'CREAT A ZLD?' && v === 'NO CONSTA') cls += ' no';
        if(c.k === 'DIF' && v < 0) cls += ' neg';
        if(c.key) cls += ' key';
        if(i===0) cls += ' sticky';
        if(v === null || v === undefined) v = '';
        else if(c.n && typeof v === 'number') v = Number.isInteger(v) ? NUMFMT.format(v) : v.toFixed(1);
        h += '<td class="'+cls+'">'+esc(v)+'</td>';
      });
      h += '</tr>';
    }
    tbody.innerHTML = h;
    let f = '';
    cols.forEach((c,i) => {
      const st = i===0 ? ' sticky' : '';
      if(c.sum){ const s = out.reduce((a,r)=>a+(Number(r[c.k])||0),0); f += '<td class="num'+st+'">'+NUMFMT.format(s)+'</td>'; }
      else f += '<td class="'+st.trim()+'">'+(i===0 ? 'TOTAL ('+NUMFMT.format(out.length)+' files)' : '')+'</td>';
    });
    tfoot.innerHTML = f;
    document.getElementById('c-'+id).textContent = out.length + ' files' + (out.length>MAX ? ' (es mostren '+MAX+')' : '');
    const k = document.getElementById('k-'+id);
    const repoRows = out.filter(r=>r.REPO>0);
    const sum = key => out.reduce((a,r)=>a+(Number(r[key])||0),0);
    k.innerHTML = spec.kpis.map(x => '<div class="kpi"><b>'+NUMFMT.format(x.k==='__rows__'?repoRows.length:sum(x.k))+'</b><span>'+esc(x.l)+'</span></div>').join('');
    fitHeight();
  }
  panel.querySelector('#q-'+id).addEventListener('input', e => { state.q = e.target.value; render(); });
  panel.querySelector('#r-'+id).addEventListener('change', e => { state.onlyRepo = e.target.checked; render(); });
  panel.querySelectorAll('select').forEach(s => s.addEventListener('change', e => { state.filters[e.target.dataset.f] = e.target.value; render(); }));
  colbtn.addEventListener('click', e => { e.stopPropagation(); colpick.hidden = !colpick.hidden; });
  colpick.addEventListener('click', e => e.stopPropagation());
  document.addEventListener('click', () => { colpick.hidden = true; });
  let syncing = false;
  topscroll.addEventListener('scroll', () => { if(syncing) return; syncing = true; wrap.scrollLeft = topscroll.scrollLeft; syncing = false; });
  wrap.addEventListener('scroll', () => { if(syncing) return; syncing = true; topscroll.scrollLeft = wrap.scrollLeft; syncing = false; });
  window.addEventListener('resize', fitHeight);
  renderPicker(); renderHead(); render();
  return { fit: fitHeight };
}
const views = { mc: build('mc', DATA.mcSpec, DATA.mc), sku: build('sku', DATA.skuSpec, DATA.sku) };
document.querySelectorAll('.tab').forEach(t => t.addEventListener('click', () => {
  document.querySelectorAll('.tab').forEach(x=>x.classList.toggle('active', x===t));
  document.querySelectorAll('.panel').forEach(p=>p.classList.toggle('active', p.id==='p-'+t.dataset.t));
  views[t.dataset.t].fit();
}));
</script></body></html>
"""


def write_html(sku: pd.DataFrame, mc: pd.DataFrame, title: str, subtitle: str, warnings: list[str], path: str):
    sku = sku.drop(columns=[c for c in HTML_HIDE if c in sku.columns])
    mc = mc.drop(columns=[c for c in HTML_HIDE if c in mc.columns])

    def recs(df):
        out, cols = [], list(df.columns)
        for row in df.itertuples(index=False, name=None):
            rec = {}
            for k, v in zip(cols, row):
                if v is None or v is pd.NA or (isinstance(v, float) and np.isnan(v)):
                    rec[k] = None
                elif isinstance(v, np.integer):
                    rec[k] = int(v)
                elif isinstance(v, np.floating):
                    rec[k] = float(v)
                else:
                    rec[k] = v
            out.append(rec)
        return out
    num_sku = {c for c in sku.columns if pd.api.types.is_numeric_dtype(sku[c])}
    num_mc = {c for c in mc.columns if pd.api.types.is_numeric_dtype(mc[c])}
    sum_cols = {"HAURIA", "STOCK ZLD", "OFFERABLE", "ENV PENDENTS", "DIF", "REPO", "PREPARABLE", "FALTA STOCK TP", "STOCK TP 01 02",
                "DISPO 30 DIES", "DISPO 59 DIES", "VENDA 1 SETM", "VENDA SKU 1 SETM", "ACUM'25", "ACUM'26", "VENDA 4 SETM"}
    key_cols = {"HAURIA", "DIF", "REPO", "PREPARABLE"}
    def spec(df, nums, default_sort, only_repo, facets, search, kpis, sum_ok, header_groups=()):
        names = list(df.columns)
        hg = {}
        for start, end, cls in header_groups:
            if start in names and end in names:
                for k in range(names.index(start), names.index(end) + 1):
                    hg[names[k]] = cls
        return {"cols": [{"k": c, "l": c, "n": c in nums, "sum": c in sum_ok, "key": c in key_cols, "hg": hg.get(c, "")} for c in df.columns],
                "defaultSort": default_sort, "onlyRepoDefault": only_repo, "facets": facets, "search": search, "kpis": kpis}
    mc_sums = sum_cols - {"VENDA 1 SETM", "VENDA 4 SETM"} | {"VENDA 1 SETM"}
    data = {
        "mc": recs(mc), "sku": recs(sku),
        "mcSpec": spec(mc, num_mc, "REPO", False, ["GÈNERE", "SEASON", "TEMPORADA", "COL·LECCIÓ", "CREAT A ZLD?"], ["model_color", "model", "color", "COL·LECCIÓ", "AVÍS"],
                       [{"k": "__rows__", "l": "model_color amb REPO"}, {"k": "REPO", "l": "parells REPO"}, {"k": "PREPARABLE", "l": "preparables (stock 59d)"},
                        {"k": "VENDA 1 SETM", "l": "venda setmana"}, {"k": "STOCK ZLD", "l": "stock Zalando"}, {"k": "ENV PENDENTS", "l": "env. pendents"}], mc_sums,
                       (("model_color", "CREAT A ZLD?", "grey"), ("VENDA 1 SETM", "OBJECTIU", "yellow"))),
        "skuSpec": spec(sku, num_sku, "REPO", True, ["GÈNERE", "SEASON", "TEMPORADA", "CREAT A ZLD?"], ["EAN", "SKU", "model_color", "model", "color", "talla", "AVÍS"],
                        [{"k": "__rows__", "l": "SKUs amb REPO"}, {"k": "REPO", "l": "parells REPO"}, {"k": "PREPARABLE", "l": "preparables (stock 59d)"},
                         {"k": "STOCK ZLD", "l": "stock Zalando"}], sum_cols - {"VENDA 1 SETM", "VENDA 4 SETM", "ACUM'25", "ACUM'26"},
                        (("EAN", "CREAT A ZLD?", "grey"), ("VENDA 1 SETM", "OBJECTIU", "yellow"))),
    }
    warn_html = "".join(f'<div class="warn">{html.escape(w)}</div>' for w in warnings)
    page = (HTML_TEMPLATE.replace("__TITLE__", html.escape(title)).replace("__SUBTITLE__", html.escape(subtitle))
            .replace("__WARNINGS__", warn_html).replace("__DATA__", json.dumps(data, ensure_ascii=False).replace("</", "<\\/")))
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(page)


# --------------------------------------------------------------------------------------
def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="Càlcul de reposició Zalando")
    ap.add_argument("--data", default=here, help="carpeta amb les fonts")
    ap.add_argument("--out", default=None, help="carpeta de sortida (per defecte la de dades)")
    ap.add_argument("--mult", type=float, default=2.0, help="multiplicador de la venda setmanal (2)")
    ap.add_argument("--min-level", type=int, default=0, help="nivell mínim per adults encara que no hi hagi venda (0 = cap)")
    ap.add_argument("--min-level-kids", type=int, default=0, help="nivell mínim per nens (0 = cap)")
    ap.add_argument("--max-level", type=int, default=None, help="nivell màxim (cap = sense límit)")
    ap.add_argument("--date", default=dt.date.today().strftime("%d.%m"), help="etiqueta de data pels fitxers de sortida (dd.mm)")
    ap.add_argument("--snapshot", action="append", default=[], help="fitxer(s) de snapshot addicional(s) a considerar")
    args = ap.parse_args()
    data, out = args.data, args.out or args.data
    os.makedirs(out, exist_ok=True)
    cache_dir = os.path.join(os.environ.get("LOCALAPPDATA", here), "Temp", "zalando_repo_cache")

    print("Llegint fonts...")
    models = load_models(os.path.join(data, "Models a reposar.xlsx"))
    levels = load_levels(os.path.join(data, "NIVEL.xlsx"))
    lines, sources = load_sales(data, cache_dir)
    acum25 = load_sales_2025(os.path.join(data, "Venda 2025.xlsx"))
    stock_tp, tp_file = load_stock_tp(os.path.join(data, "Stock Toni Pons"))
    snap, snap_meta = load_snapshot(os.path.join(data, "Stock Zalando"), args.snapshot)
    pending, pend_labels = load_pending(os.path.join(data, "Enviaments pendents"))
    adjust = load_adjustments(os.path.join(data, "Ajustos repo.xlsx"))

    print("Calculant...")
    sku, mc, fora, info = compute(models, levels, lines, acum25, stock_tp, snap, pending, pend_labels, adjust,
                                  args.mult, args.min_level, args.min_level_kids, args.max_level)

    warnings = []
    if snap_meta["rebutjats"]:
        warnings.append("Snapshots de stock Zalando rebutjats per EANs no íntegres (notació científica 8,43453E+12 en desar des d'Excel): "
                        + ", ".join(snap_meta["rebutjats"]) + ". Cal desar el CSV original de Zalando sense obrir-lo amb Excel.")
    snap_date = dt.date.fromisoformat(snap_meta["data"])
    if (dt.date.today() - snap_date).days > 3:
        warnings.append(f"El snapshot de stock Zalando utilitzat és del {snap_date.strftime('%d.%m.%Y')} ({(dt.date.today()-snap_date).days} dies).")
    no_created = int((mc["CREAT A ZLD?"] == "NO CONSTA").sum())
    unmatched_env = int(pending.loc[~pending["EAN"].isin(set(models["EAN"].dropna())), "ENV PENDENTS"].sum()) if len(pending) else 0
    if unmatched_env:
        warnings.append(f"{unmatched_env} parells dels enviaments pendents tenen EANs que no són a 'Models a reposar' (no es resten enlloc).")
    if len(adjust):
        warnings.append(f"S'han aplicat {len(adjust)} ajustos de 'Ajustos repo.xlsx' (multiplicador / nivell forçat).")

    week_lbl = f"{info['setmana_inici'][8:10]}.{info['setmana_inici'][5:7]} - {info['setmana_fi'][8:10]}.{info['setmana_fi'][5:7]}"
    params = [
        ("Data càlcul", dt.date.today().isoformat()),
        ("Setmana de venda utilitzada", f"{week_lbl} ({info['setmanes']} setmanes acumulades el 2026)"),
        ("Regla", f"objectiu = venda setmanal del model_color x {args.mult:g}; nivell = primer nivell de la taula del gènere el total de parells del qual (totes les talles) cobreix l'objectiu; HAURIA = desglossament per talla d'aquest nivell"),
        ("Nivell mínim adults / nens", f"{args.min_level} / {args.min_level_kids} (0 = sense venda no es reposa)"),
        ("Nivell màxim", str(args.max_level) if args.max_level else "sense límit (el de la taula)"),
        ("DIF", "HAURIA - STOCK ZLD (total, offerable + non-offerable) - ENV PENDENTS"),
        ("REPO", "DIF si > 0; els HI26 NOU sense marca a 'es pot enviar?' (CREAT A ZLD? = NO CONSTA) es deixen a 0 i es veuen a DIF"),
        ("PREPARABLE", "min(REPO, Stock Disponible 59 Dies a Toni Pons); FALTA STOCK TP = REPO - PREPARABLE"),
        ("Gèneres -> taula de nivells", "DONA, UNISEX -> MUJER (unisex 46-47 = talla 45) | HOME -> CABALLERO | NENS, MINI -> NIÑO (mini <25 = talla 25) | COMPLEMENTS -> objectiu directe"),
        ("Models a reposar", f"{len(models)} SKUs, {models['model_color'].nunique()} model_color"),
        ("Stock Zalando", f"{snap_meta['fitxer']} ({snap_meta['data']}), {snap_meta['eans']} EANs, {snap_meta['total']} parells"),
        ("Snapshots rebutjats", ", ".join(snap_meta["rebutjats"]) or "-"),
        ("Enviaments pendents", ", ".join(pend_labels) + f" = {int(pending['ENV PENDENTS'].sum()) if len(pending) else 0} parells"),
        ("Stock Toni Pons", tp_file),
        ("Vendes 2026", f"{len(sources)} fitxers setmanals (pestanya DADES2), {int(lines['units'].sum())} unitats"),
        ("Venda 2025", f"{int(acum25.sum())} unitats, {len(acum25)} model_color"),
        ("Model_color HI26 NOU que no consten creats a ZLD (REPO = 0)", str(no_created)),
        ("Avisos", " | ".join(warnings) or "-"),
    ]

    print("Escrivint sortides...")
    venda_xlsx = os.path.join(out, f"Venda 2026 {args.date}.xlsx")
    write_vendes(lines, sources, acum25, venda_xlsx, info)
    xlsx = os.path.join(out, f"REPO ZALANDO {args.date}.xlsx")
    write_excel(sku, mc, fora, params, levels, xlsx)
    title = f"Reposició Zalando {args.date}"
    subtitle = (f"Setmana {week_lbl} · stock Zalando {snap_meta['fitxer']} · enviaments pendents {', '.join(pend_labels) or 'cap'} · "
                f"regla venda x{args.mult:g} → nivell · generat {dt.datetime.now():%d/%m/%Y %H:%M}")
    write_html(sku, mc, title, subtitle, warnings, os.path.join(out, f"REPO ZALANDO {args.date}.html"))

    # resum
    print()
    print(f"Setmana: {week_lbl}   Snapshot: {snap_meta['fitxer']}   Pendents: {pend_labels}")
    print(f"SKUs: {len(sku)}   model_color: {len(mc)}   model_color amb venda: {(mc['VENDA 1 SETM']>0).sum()}   amb REPO: {(mc['REPO']>0).sum()}")
    print(f"REPO total: {int(sku['REPO'].sum())}   PREPARABLE: {int(sku['PREPARABLE'].sum())}   FALTA STOCK TP: {int((sku['REPO'] - sku['PREPARABLE']).sum())}")
    for w in warnings:
        print("AVÍS:", w)
    print("Sortides:", venda_xlsx, "|", xlsx, "|", xlsx[:-5] + ".html")


if __name__ == "__main__":
    main()
